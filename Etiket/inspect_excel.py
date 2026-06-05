import openpyxl, os
files = [
    r'C:\projects\Etiket\2026.05.09 KESM.XLSX',
    r'C:\projects\Etiket\12.05.2026kesme listesi - Kopya.xlsx',
    r'C:\projects\Etiket\14.05zen amorf,lıza,nora listesi.xlsx',
]
for f in files:
    print('\nFILE:', f)
    print('EXISTS:', os.path.exists(f))
    if not os.path.exists(f):
        continue
    wb = openpyxl.load_workbook(f, data_only=True)
    print(' sheets:', wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1,c).value for c in range(1,16)]
    print('HEADER:', headers)
    print('LONG Satir rows:')
    for r in range(2, min(ws.max_row, 120) + 1):
        satir = ws.cell(r, 5).value
        if isinstance(satir, str) and len(satir) > 60:
            malz = ws.cell(r, 4).value
            bek = ws.cell(r, 7).value
            print('ROW', r, 'len', len(satir), 'BEK=', bek)
            print('  MALZ:', malz)
            print('  SATIR:', satir)
    print('---')
